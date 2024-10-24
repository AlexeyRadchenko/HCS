#from argparse import ArgumentParser

from api.utils.utils import init_mkd_works_db_data, init_mkd_works_db_works_reference_book
from asyncio import get_event_loop
from openpyxl import load_workbook


def load_data_from_xlsx(filename):
    wb = load_workbook(filename=filename)
    sheet = wb.active
    data_list = []
    
    for index, row in enumerate(sheet.rows):
        if index > 0:
            data_list.append({
                'street': row[0].value,
                'number': str(row[1].value),
                'director': row[2].value,
                'director_appartment': row[3].value
            })
    return data_list

def load_data_from_sprav_xlsx(filename):
    wb = load_workbook(filename=filename)
    sheet = wb.active
    mainworks_list = []
    subworks_list = []
    fixworks_list = []
    sub_work_num = ''
    fix_work_num = ''
    for index, row in enumerate(sheet.rows):
        if row[0].value and len(row[0].value) <= 3:
            mainworks_list.append((row[0].value, row[1].value))
            sub_work_num = ''
            fix_work_num = ''
            continue
        if row[0].value and len(row[0].value) > 3 and row[0].value[-2] == '1':
            subworks_list.append((row[0].value, row[1].value, row[3].value, row[4].value))
            sub_work_num = row[0].value
            continue
        if row[0].value and len(row[0].value) > 3 and row[0].value[-2] == '2':
            fix_work_num = row[0].value
            sub_work_num = ''
            continue
        if not row[0].value and row[1].value and sub_work_num != '' and fix_work_num == '':
            subworks_list.append((sub_work_num, row[1].value, row[2].value, row[3].value, row[4].value))
            continue
        if not row[0].value and row[1].value and sub_work_num == '' and fix_work_num != '':
            fixworks_list.append((fix_work_num, row[1].value, row[2].value, row[3].value, row[4].value))

    return mainworks_list, subworks_list, fixworks_list   

if __name__ == '__main__':
    mainwork_data, subwork_data, fixwork_data = load_data_from_sprav_xlsx('works_init_sprav.xlsx')
    """print("DON'T FORGET DEL DATE_CREATE FOR SQLITE IN INIT_CONTACTS_DB_DATA")
    data_komf = load_data_from_xlsx('komf_houses.xlsx')
    data_jks = load_data_from_xlsx('jks_houses.xlsx')
    print(len(data_komf), data_komf[0])
    print(len(data_jks), data_jks[0])"""
    loop = get_event_loop()
    """
    loop.run_until_complete(init_mkd_works_db_data(data_komf, 1))
    loop.run_until_complete(init_mkd_works_db_data(data_jks, 2))"""
    loop.run_until_complete(init_mkd_works_db_works_reference_book(mainwork_data, subwork_data, fixwork_data))
            

    """parser = ArgumentParser(description="Create superuser for first start service")
    parser.add_argument("--login", "-l", action="store", dest='login', help="Enter superuser login", required=True)
    parser.add_argument("--password", "-p", action="store", dest='password', help="Enter superuser password", required=True)
    args = parser.parse_args()
    if args.login and args.password:
        loop = get_event_loop()
        loop.run_until_complete(init_superuser(args.login, args.password))
    if not args.password:
        print("Enter superuser password")
    if not args.login:
        print("Enter superuser login")"""
